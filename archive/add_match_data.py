#!/usr/bin/env python3
"""
Quick Match Data Entry Script
Reads data from the 'Data Entry' sheet and adds it to the appropriate sheets
"""

import openpyxl
from datetime import datetime
import sys

def add_match_data(excel_file):
    """Transfer data from Data Entry sheet to Matches, Goals, and Appearances sheets"""

    wb = openpyxl.load_workbook(excel_file)

    # Read from Data Entry sheet
    entry = wb['Data Entry']

    # Extract match details
    match_id = entry['B6'].value
    date = entry['B7'].value
    opponent = entry['B8'].value
    home_away = entry['B9'].value
    competition = entry['B10'].value
    team_goals = entry['B11'].value
    opp_goals = entry['B12'].value
    potm = entry['B13'].value

    # Validate required fields
    if not all([match_id, date, opponent, home_away, team_goals is not None, opp_goals is not None]):
        print("❌ Error: Please fill in all match details (yellow cells in section 1)")
        return False

    # Calculate result
    if team_goals > opp_goals:
        result = 'W'
    elif team_goals < opp_goals:
        result = 'L'
    else:
        result = 'D'

    # Add to Matches sheet
    matches = wb['Matches']
    next_row = matches.max_row + 1

    matches[f'A{next_row}'] = match_id
    matches[f'B{next_row}'] = date
    matches[f'C{next_row}'] = opponent
    matches[f'D{next_row}'] = home_away
    matches[f'E{next_row}'] = competition
    matches[f'F{next_row}'] = team_goals
    matches[f'G{next_row}'] = opp_goals
    matches[f'H{next_row}'] = result
    matches[f'I{next_row}'] = potm
    matches[f'J{next_row}'] = f'=F{next_row}-G{next_row}'

    print(f"✓ Added match: {match_id} vs {opponent} ({result} {team_goals}-{opp_goals})")

    # Add goals to Goals sheet
    goals = wb['Goals']
    goals_added = 0

    for row in range(17, 27):  # Goal entry rows (up to 10 goals)
        scorer = entry.cell(row, 1).value
        assister = entry.cell(row, 2).value
        own_goal = entry.cell(row, 3).value

        if scorer:  # Only add if scorer is filled in
            next_goal_row = goals.max_row + 1
            goals[f'A{next_goal_row}'] = match_id
            goals[f'B{next_goal_row}'] = date
            goals[f'C{next_goal_row}'] = scorer
            goals[f'D{next_goal_row}'] = assister if assister else ''
            goals[f'E{next_goal_row}'] = own_goal if own_goal else 'No'
            goals_added += 1

    print(f"✓ Added {goals_added} goals")

    # Add appearances to Appearances sheet
    appearances = wb['Appearances']
    players = ['Fabian', 'Lake', 'Darius', 'Siji', 'Reuben', 'Kian', 'Aydan', 'Leo']

    # Find or add column for this match
    match_col = None
    for col in range(2, 10):  # Check columns B-I
        if appearances.cell(1, col).value == match_id:
            match_col = col
            break

    if not match_col:
        # Add new column for this match
        match_col = appearances.max_column + 1
        appearances.cell(1, match_col, match_id)
        print(f"✓ Added new match column: {match_id}")

    # Update appearances
    appearances_count = 0
    for idx, player in enumerate(players):
        played = entry.cell(33 + idx, 2).value  # Starting at row 33 for player appearances

        # Find player row in Appearances sheet
        for app_row in range(2, 10):
            if appearances.cell(app_row, 1).value == player:
                appearances.cell(app_row, match_col, played if played else 0)
                if played == 1:
                    appearances_count += 1
                break

    print(f"✓ Recorded {appearances_count} player appearances")

    # Save workbook
    wb.save(excel_file)
    print(f"\n✅ Match data successfully added to spreadsheet!")
    print(f"📊 Dashboard and stats will update automatically")

    # Clear the data entry form
    clear = input("\nClear the Data Entry form for next match? (y/n): ")
    if clear.lower() == 'y':
        clear_entry_form(wb, excel_file)

    return True

def clear_entry_form(wb, excel_file):
    """Clear the Data Entry form after successful data transfer"""
    entry = wb['Data Entry']

    # Clear match details
    for row in [6, 7, 8, 9, 11, 12, 13]:
        entry[f'B{row}'].value = None

    entry['B10'].value = 'League'  # Reset competition to default

    # Clear goals
    for row in range(17, 27):
        for col in range(1, 4):
            if col == 3:
                entry.cell(row, col).value = 'No'  # Reset Own Goal to No
            else:
                entry.cell(row, col).value = None

    # Clear appearances
    for row in range(33, 41):
        entry.cell(row, 2).value = None

    wb.save(excel_file)
    print("✓ Data Entry form cleared and ready for next match")

if __name__ == "__main__":
    excel_file = "Aspley_Guise_Rhinos_Enhanced.xlsx"

    print("="*60)
    print("ASPLEY GUISE RHINOS U8s - QUICK DATA ENTRY")
    print("="*60)
    print(f"\nReading from: {excel_file}")
    print("Make sure you've filled in the 'Data Entry' sheet first!\n")

    try:
        add_match_data(excel_file)
    except FileNotFoundError:
        print(f"❌ Error: Could not find {excel_file}")
        print("Make sure this script is in the same folder as your spreadsheet")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
